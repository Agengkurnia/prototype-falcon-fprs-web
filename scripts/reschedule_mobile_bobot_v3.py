#!/usr/bin/env python3
"""Reschedule MOBILE track in bobot_kesulitan_modul_fprs_v3.xlsx from today (2026-08-10)."""
from __future__ import annotations

from datetime import date, datetime, timedelta

import openpyxl

PATH = r"D:\Work\Source\Comsup\falcon\Prototype\docs\bobot_kesulitan_modul_fprs_v3.xlsx"
HOLIDAYS = {date(2026, 8, 17), date(2026, 8, 25)}
START = date(2026, 8, 10)
DATE_FMT = "dd-mmm-yyyy"
ID_MON = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]


def is_work(d: date) -> bool:
    return d.weekday() < 5 and d not in HOLIDAYS


def allocate(start_from: date, needed: int) -> tuple[date, date]:
    d = start_from
    while not is_work(d):
        d += timedelta(days=1)
    start = d
    end = d
    count = 0
    while count < needed:
        if is_work(end):
            count += 1
        if count < needed:
            end += timedelta(days=1)
    return start, end


def set_date(cell, d: date) -> None:
    cell.value = datetime(d.year, d.month, d.day)
    cell.number_format = DATE_FMT


def fmt_id(d: date) -> str:
    return f"{d.day} {ID_MON[d.month - 1]} {d.year}"


def main() -> None:
    wb = openpyxl.load_workbook(PATH)
    ws = wb["Timeline MOBILE"]

    cursor = START
    first = None
    last = None
    rows_out: list[tuple[str, date, date, int]] = []

    for r in range(2, ws.max_row + 1):
        modul = ws.cell(r, 3).value
        if not modul or str(modul).strip().upper() == "TOTAL":
            continue
        hari = int(ws.cell(r, 7).value or 0)
        s, e = allocate(cursor, hari)
        set_date(ws.cell(r, 5), s)
        set_date(ws.cell(r, 6), e)
        rows_out.append((str(modul), s, e, hari))
        if first is None:
            first = s
        last = e
        cursor = e + timedelta(days=1)
        print(f"{r:2d} {hari}d  {s} -> {e}  {modul}")

    assert first and last
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 3).value or "").strip().upper() != "TOTAL":
            continue
        set_date(ws.cell(r, 5), first)
        set_date(ws.cell(r, 6), last)
        ws.cell(r, 8).value = (
            f"Timeline MOBILE | excl. weekend + libur nasional RI | "
            f"{fmt_id(first)} – {fmt_id(last)} | "
            f"v3: reschedule start {START.isoformat()} (+4 FALCON MERGER)"
        )
        print(f"TOTAL {first} -> {last}")
        break

    by_modul = {m: (s, e) for m, s, e, _ in rows_out}
    ws2 = wb["Semua Modul"]
    for r in range(2, ws2.max_row + 1):
        if str(ws2.cell(r, 2).value or "").upper() != "MOBILE":
            continue
        modul = str(ws2.cell(r, 3).value or "")
        if modul not in by_modul:
            print("WARN missing", modul)
            continue
        s, e = by_modul[modul]
        set_date(ws2.cell(r, 5), s)
        set_date(ws2.cell(r, 6), e)

    ws3 = wb["Ringkasan"]
    ws3.cell(2, 1).value = (
        f"WEB dan MOBILE paralel (dev berbeda). WEB: 22 Jul – 30 Sep 2026. "
        f"MOBILE: {fmt_id(first)} – {fmt_id(last)}. "
        f"Hari kerja = Senin–Jumat dikurangi libur nasional/cuti bersama RI (SKB 3 Menteri 2026). "
        f"v3: hapus Canvassing (take-out); +4 FALCON MERGER; "
        f"Mobile di-reschedule mulai {fmt_id(START)}."
    )

    for r in range(4, ws3.max_row + 1):
        plat = str(ws3.cell(r, 1).value or "").upper()
        if plat == "MOBILE":
            set_date(ws3.cell(r, 4), first)
            set_date(ws3.cell(r, 5), last)
        elif plat == "TOTAL":
            set_date(ws3.cell(r, 4), date(2026, 7, 22))
            set_date(ws3.cell(r, 5), last)
            ws3.cell(r, 6).value = "43 + 92 (2 track)"

    for r in range(1, ws3.max_row + 1):
        v = str(ws3.cell(r, 1).value or "")
        if v.startswith("4. Track WEB"):
            ws3.cell(r, 1).value = (
                f"4. Track WEB: 22-Jul-2026 s/d 30-Sep-2026 (Canvassing dihapus di v3). "
                f"Track MOBILE: {fmt_id(START)} s/d {fmt_id(last)} "
                f"(reschedule {START.isoformat()}; + FALCON MERGER setelah Sinkronisasi)."
            )

    wb.save(PATH)
    total_hari = sum(h for _, _, _, h in rows_out)
    print(f"Saved {PATH}")
    print(f"Mobile: {first} -> {last} ({total_hari} hari kerja, {len(rows_out)} modul)")


if __name__ == "__main__":
    main()
