"""Generate pegawai.json from 'Master Akun Simplidots.xlsx'.

Roles are separated by sheet:
  - USER (MOTORIS)  -> role "Motoris"   (kolom: NO, Employee Code (NIK), Name, STATUS USER, STATUS)
  - USER (SPG GT)   -> role "SPG GT"    (kolom: NO, NIK, NAMA LENGKAP, NO TLP, BRANCH)

Region diturunkan dari BRANCH (grouping geografis, konsisten dgn Master Stokis).
"""
import json, os
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, 'master-akun-simplidots.xlsx')
OUT = os.path.join(BASE, 'pegawai.json')

# Branch (kota) -> Region. Konsisten dengan Master Stokis (Sumatera=1, Jabodetabek=2, Jabar=3).
BRANCH_REGION = {
    # Region 1 — Sumatera
    'ACEH': 'Region 1', 'MEDAN': 'Region 1', 'PEMATANG SIANTAR': 'Region 1',
    'PADANG': 'Region 1', 'PEKANBARU': 'Region 1', 'BATAM': 'Region 1',
    'JAMBI': 'Region 1', 'PALEMBANG': 'Region 1', 'BENGKULU': 'Region 1',
    'LAMPUNG': 'Region 1', 'PANGKAL PINANG': 'Region 1',
    # Region 2 — Jabodetabek & Banten
    'JAKARTA 1': 'Region 2', 'BEKASI': 'Region 2', 'BOGOR': 'Region 2',
    'DEPOK': 'Region 2', 'TANGERANG': 'Region 2',
    # Region 3 — Jawa Barat
    'BANDUNG': 'Region 3', 'CIREBON': 'Region 3', 'TASIKMALAYA': 'Region 3',
    # Region 4 — Jawa Tengah & DIY
    'SEMARANG': 'Region 4', 'SOLO': 'Region 4', 'KUDUS': 'Region 4',
    'TEGAL': 'Region 4', 'PURWOKERTO': 'Region 4', 'YOGYAKARTA': 'Region 4',
    # Region 5 — Jawa Timur
    'SURABAYA1': 'Region 5', 'SURABAYA2': 'Region 5', 'MALANG': 'Region 5',
    'KEDIRI': 'Region 5', 'JEMBER': 'Region 5',
    # Region 6 — Bali & Nusa Tenggara
    'DENPASAR': 'Region 6', 'MATARAM': 'Region 6', 'KUPANG': 'Region 6',
    # Region 7 — Kalimantan
    'BALIKPAPAN': 'Region 7', 'BANJARMASIN': 'Region 7', 'PONTIANAK': 'Region 7',
    'SAMARINDA': 'Region 7',
    # Region 8 — Sulawesi
    'MAKASSAR': 'Region 8', 'PALU': 'Region 8',
}


def norm_phone(val):
    if val in (None, '', '-'):
        return ''
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    s = ''.join(ch for ch in s if ch.isdigit())
    if not s:
        return ''
    if not s.startswith('0'):
        s = '0' + s
    return s


def norm_code(val):
    if val in (None, '', '-'):
        return ''
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s


def title_branch(b):
    if not b:
        return ''
    return str(b).strip().title().replace('Spg', 'SPG')


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    out = []
    nid = 1

    # Baris yang di-skip: slot kosong / dummy
    SKIP_NAMA = {'VACANT', 'BLORA1'}

    # --- Motoris ---
    ws = wb['USER (MOTORIS)']
    for r in ws.iter_rows(min_row=2, values_only=True):
        nama = (r[2] or '').strip() if r[2] else ''
        if not nama or nama.upper() in SKIP_NAMA:
            continue
        status_user = (r[3] or '').strip() if r[3] else ''
        status = (r[4] or '').strip().title() if r[4] else 'Active'
        out.append({
            'id': nid,
            'kode': norm_code(r[1]),
            'nama': nama,
            'role': 'Motoris',
            'telepon': '',
            'branch': '',
            'region': '',
            'keterangan': status_user,
            'status': 'Active' if status.upper() == 'ACTIVE' else (status or 'Active'),
        })
        nid += 1

    # --- SPG GT ---
    ws = wb['USER (SPG GT)']
    for r in ws.iter_rows(min_row=2, values_only=True):
        nama = (r[2] or '').strip() if r[2] else ''
        if not nama:
            continue
        branch_raw = (str(r[4]).strip().upper() if r[4] else '')
        region = BRANCH_REGION.get(branch_raw, '')
        out.append({
            'id': nid,
            'kode': norm_code(r[1]),
            'nama': nama,
            'role': 'SPG GT',
            'telepon': norm_phone(r[3]),
            'branch': title_branch(r[4]),
            'region': region,
            'keterangan': '',
            'status': 'Active',
        })
        nid += 1

    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    from collections import Counter
    print(f'Total pegawai: {len(out)}')
    print('Per role:', dict(Counter(p['role'] for p in out)))
    print('Per region:', dict(Counter(p['region'] or '(kosong)' for p in out)))
    missing = sorted(set(p['branch'] for p in out if p['role'] == 'SPG GT' and not p['region']))
    if missing:
        print('BRANCH tanpa region (perlu dipetakan):', missing)


if __name__ == '__main__':
    main()
